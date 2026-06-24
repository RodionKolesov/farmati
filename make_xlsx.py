# -*- coding: utf-8 -*-
"""Сравнительная таблица вариантов сайта -> Excel (.xlsx) в фирменном стиле FARMATI."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PLUM = "6C4D7A"
PLUM_D = "573E63"
BLUSH = "F8EBEE"
BLUSH2 = "FDF7F8"
INK = "3A2E3A"
LINE = "E3D7DD"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Сравнение"

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="center")
wrap_c = Alignment(wrap_text=True, vertical="center", horizontal="center")

# Заголовок-шапка
ws.merge_cells("A1:C1")
c = ws["A1"]
c.value = "Сайт «Формула красоты» — сравнение вариантов"
c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
c.fill = PatternFill("solid", fgColor=PLUM_D)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Шапка колонок
headers = ["Критерий", "Вариант А — готовые сервисы\n(Tilda + сервис бонусов)", "Вариант Б — свой сайт\n(с базой данных)"]
for i, h in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=i, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=PLUM)
    cell.alignment = wrap_c
    cell.border = border
ws.row_dimensions[2].height = 42

rows = [
    ("Что это", "Сайт на конструкторе Tilda + готовый сервис лояльности", "Сайт, написанный под вас, со своей базой данных"),
    ("Оплата в месяц", "~2 300–3 300 ₽ (Tilda + сервис бонусов)", "~600–2 000 ₽ (сервер + домен)"),
    ("Разовые затраты", "Минимальные (настройка)", "Заметные (разработка + панель управления)"),
    ("Кто ведёт товары и контент", "Вы сами — мышкой в простой панели", "Через панель управления (её нужно сделать)"),
    ("Нужен ли программист потом", "Нет", "Да — для поддержки и доработок"),
    ("Бонусы и личный кабинет", "Готовые, «из коробки»", "Свои, максимально гибкие"),
    ("Оплата бонусами в корзине", "Да (через сервис)", "Да, прямо в корзине"),
    ("Надёжность / «само работает»", "Высокая — за работу отвечают сервисы", "Зависит от программиста и сервера"),
    ("Гибкость дизайна", "Высокая, но в рамках Tilda", "Полная — любой дизайн"),
    ("Срок запуска", "Быстро (дни)", "Дольше (недели)"),
    ("Кому подходит", "«Хочу запустить и вести сама, без техники»", "«Готова держать программиста, нужен контроль»"),
]

r = 3
for label, a, b in rows:
    lc = ws.cell(row=r, column=1, value=label)
    lc.font = Font(bold=True, color=INK)
    lc.fill = PatternFill("solid", fgColor=BLUSH)
    lc.alignment = wrap
    lc.border = border
    for col, val in ((2, a), (3, b)):
        cc = ws.cell(row=r, column=col, value=val)
        cc.alignment = wrap
        cc.border = border
        cc.fill = PatternFill("solid", fgColor=(BLUSH2 if r % 2 else "FFFFFF"))
        cc.font = Font(color=INK)
    ws.row_dimensions[r].height = 34
    r += 1

# Итоговая рекомендация
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
rec = ws.cell(row=r, column=1,
    value="Рекомендация: хотите «отдали и веду сама» → Вариант А. Есть постоянный программист и нужен полный контроль → Вариант Б. Бонусы можно подключить сразу или вторым этапом.")
rec.font = Font(italic=True, color=PLUM_D)
rec.alignment = Alignment(wrap_text=True, vertical="center")
rec.fill = PatternFill("solid", fgColor=BLUSH)
rec.border = border
ws.row_dimensions[r].height = 46

ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 42
ws.column_dimensions["C"].width = 42
ws.freeze_panes = "A3"

out = "Сравнение-вариантов.xlsx"
wb.save(out)
print("Сохранено:", out)
